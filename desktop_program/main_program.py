# -*- coding: utf-8 -*-
import mysql.connector
from mysql.connector import Error
from docxtpl import DocxTemplate
from datetime import date
import sys
import os
import tkinter as tk
from openpyxl import *
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog as fd
import urllib.request


def import_ancet():
    file_name = fd.askopenfilename()
    wb = load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    stud_name = (sheet.cell(row=2, column=2).value)
    stud_sur = (sheet.cell(row=3, column=2).value)
    stud_otch = (sheet.cell(row=4, column=2).value)
    stud_age = str((sheet.cell(row=5, column=2).value))
    par_name = (sheet.cell(row=6, column=2).value)
    par_sur = str((sheet.cell(row=7, column=2).value))
    par_otch = (sheet.cell(row=8, column=2).value)
    par_ser = (sheet.cell(row=11, column=2).value)
    par_pasnum = (sheet.cell(row=12, column=2).value)
    par_vidan = (sheet.cell(row=13, column=2).value)
    par_num = (sheet.cell(row=9, column=2).value)
    lesson = (sheet.cell(row=10, column=2).value)
    wb.close()

    cursor.execute(f"insert into parent values (default,'{par_name}', '{par_sur}', '{par_otch}', '{par_num}', '{par_ser}', '{par_pasnum}', '{par_vidan}')")
    connection.commit()
    cursor.execute("select id from parent where id = (select max(id) from parent)")
    parent_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"insert into students values (default,'{stud_name}', '{stud_sur}', '{stud_otch}', '{stud_age}', '{parent_id}', '0', '0')")
    connection.commit()
    cursor.execute("select id from students where id = (select max(id) from students)")
    student_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"select id from subjects where name = '{lesson}'")
    subject_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"select Teacher_id from subjects where name = '{lesson}'")
    teacher_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"insert into students_has_subjects values ('{student_id}', '{subject_id}')")
    connection.commit()

def check_files_dir():
    global files
    dirname = './files'
    dirfiles = os.listdir(dirname)
    fullpaths = map(lambda name: os.path.join(dirname, name), dirfiles)
    files = []
    for file in fullpaths:
        if os.path.isfile(file): files.append(file)
        
def acc_student():
    cursor.execute(f"insert into parent values (default,'{par_name}', '{par_sur}', '{par_otch}', '{par_num}', '{par_ser}', '{par_pasnum}', '{par_vidan}')")
    connection.commit()
    cursor.execute("select id from parent where id = (select max(id) from parent)")
    parent_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"insert into students values (default,'{stud_name}', '{stud_sur}', '{stud_otch}', '{stud_age}', '{parent_id}', '0', '0')")
    connection.commit()
    cursor.execute("select id from students where id = (select max(id) from students)")
    student_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"select id from subjects where name = '{lesson}'")
    subject_id = str(cursor.fetchall()[0][0])
    cursor.execute(f"insert into students_has_subjects values ('{student_id}', '{subject_id}')")
    connection.commit()
    
    if len(files) == 0:
        win_list.destroy()
        messagebox.showerror("Внимание!", "Заявок больше нет")
    else:
        update_list_win()

def kick_student():
    if len(files) == 0:
        win_list.destroy()
        messagebox.showinfo("Внимание!", "Заявок больше нет")
    else:
        update_list_win()
    
def update_list_win():
    global stud_name
    global stud_sur
    global stud_otch
    global stud_age
    global par_name
    global par_sur
    global par_otch
    global par_ser
    global par_pasnum
    global par_vidan
    global par_num
    global lesson
    
    text_info.delete("1.0","end")
    file_name = files[0]
    wb = load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    stud_name = (sheet.cell(row=2, column=2).value)
    stud_sur = (sheet.cell(row=3, column=2).value)
    stud_otch = (sheet.cell(row=4, column=2).value)
    stud_age = str((sheet.cell(row=5, column=2).value))
    par_name = (sheet.cell(row=6, column=2).value)
    par_sur = str((sheet.cell(row=7, column=2).value))
    par_otch = (sheet.cell(row=8, column=2).value)
    par_ser = (sheet.cell(row=11, column=2).value)
    par_pasnum = (sheet.cell(row=12, column=2).value)
    par_vidan = (sheet.cell(row=13, column=2).value)
    par_num = (sheet.cell(row=9, column=2).value)
    lesson = (sheet.cell(row=10, column=2).value)
    wb.close()
    os.remove(files[0])
    files.pop(0)
    
    message = str(f'''Имя: {stud_name}\nФамилия: {stud_sur}\nОтчество: {stud_otch}\nВозраст: {stud_age}
\nИмя родителя: {par_name}\nФамилия родителя: {par_sur}\nОтчество родителя: {par_otch}
Номер: {par_num}\n\nПредметы: {lesson}''')
    text_info.insert(1.0, message)
    

# Окно заявок работает следующим образом: проверяются заявки выгруженные на сервер при помощи функции check_files_dir.
# Затем мы подтягиваем информацию update_list_win, попутно удаляя файл заявки, следующий элемент становится первым 
# на очереди, принятие записывает информацию в БД, отказ просто переносит нас на следующую анкету при помощи update_list_win

def list_win_gui():
    global text_info
    global win_list
    check_files_dir()
    if files == []:
        messagebox.showinfo("Внимание!", "Заявок пока нет")
    else:
        win_list = tk.Toplevel(win)
        win_list.title("Ancet Loop")
        win_list.iconbitmap('icon.ico')
        win_list.geometry("350x400+400+150")
        win_list.resizable(False,False)
        
        frame3 = tk.Frame(win_list, width=350, height=400)
        btn_accept = tk.Button(frame3, text = "Принять ученика", font = ("Calibri", 12, "bold"), 
                               bg='green', fg = 'white', command = acc_student)
        btn_not_accept = tk.Button(frame3, text = "Не принимать", font = ("Calibri", 12, "bold"), 
                               bg='red', fg = 'white', command = kick_student)
        text_info = tk.Text(frame3)
        
        frame3.pack()
        btn_accept.place(x=190, y=360, height=30, width=150)
        btn_not_accept.place(x=10, y=360, height=30, width=150)
        text_info.place(x=10, y=10, height=340, width=330)
        
        update_list_win()
    
def donothing():
    print('Done')
    
def export():
    if st_name == 0:
        messagebox.showerror("Ошибка!", "Для изменения необходимо выбрать ученика")
    else:
        current_date = date.today()
        filename = fd.askopenfilename()
        doc = DocxTemplate(filename)
        context = {'pn' : pr_name, 'psur' : pr_surname, 'potch' : pr_otchestvo, 'stn' : st_name, 
                   'stsur' : st_surname, 'stotch' : st_otchestvo, 'tn' : tch_name, 'tsur' : tch_surname,
                   'totch' : tch_otchestvo, 'telephone' : pr_number, 'date' : current_date, 
                   'tser' : tch_pas_ser, 'tnum' : tch_pas_num, 'tvid' : tch_pas_vid,
                   'pser' : pr_pas_ser, 'pnum' : pr_pas_num, 'pvid' : pr_pas_vid}
        doc.render(context)
        save_name = fd.asksaveasfilename(filetypes=(("Word - документ", "*.docx"),
                                                    ("All files", "*.*")))
        doc.save(save_name+'.docx')


def search_button():
    global st_name
    global st_surname
    global st_otchestvo
    global pr_name
    global pr_surname
    global pr_otchestvo
    global pr_number
    global tch_name
    global tch_surname
    global tch_otchestvo
    global pr_pas_ser
    global pr_pas_num
    global pr_pas_vid
    global tch_pas_ser
    global tch_pas_num
    global tch_pas_vid

    try:
        textbox.delete("1.0","end")
        bd_s_name = s_name.get()
        bd_s_surname = s_surname.get()
        cursor.execute(f"select * from students where name = '{bd_s_name}' and surname = '{bd_s_surname}'")
        info = cursor.fetchall()
        st_id= str(info[0][0])
        st_name = info[0][1]
        st_surname = info[0][2]
        st_otchestvo = info[0][3]
        st_age = str(info[0][4])
        st_parent_id = str(info[0][5])
                
        cursor.execute(f"select * from parent where id = '{st_parent_id}'")
        info2 = cursor.fetchall()
        pr_name = info2[0][1]
        pr_surname = info2[0][2]
        pr_otchestvo = info2[0][3]
        pr_number = str(info2[0][4])
        pr_pas_ser = str(info2[0][5])
        pr_pas_num = str(info2[0][6])
        pr_pas_vid = str(info2[0][7])
            
        cursor.execute(f"select subjects_id from students_has_subjects where students_id = '{st_id}'")
        lesson_id = str(cursor.fetchall()[0][0])
        cursor.execute(f"select * from subjects where id = '{lesson_id}'")
        info3 = cursor.fetchall()
        teacher_id = info3[0][2]
        subjects = info3[0][1]
                                                                           
        cursor.execute(f"select * from teacher where id = '{teacher_id}'")
        info4 = cursor.fetchall()
        tch_name = info4[0][1]
        tch_surname = info4[0][2]
        tch_otchestvo = info4[0][3]
        tch_pas_ser = info4[0][6]
        tch_pas_num = info4[0][7]
        tch_pas_vid = info4[0][8]
            
        message = str(f'''Имя: {st_name}\nФамилия: {st_surname}\nОтчество: {st_otchestvo}\nВозраст: {st_age}
\nИмя родителя: {pr_name}\nФамилия родителя: {pr_surname}\nОтчество родителя: {pr_otchestvo}
Номер: {pr_number}\n\nПредметы: {subjects}\n\nУчитель: {tch_surname} {tch_name} {tch_otchestvo}''')
        textbox.insert(1.0, message)
    except:
        textbox.insert(1.0, '\nУченик не найден!')  

def show_info_about_programm():
    messagebox.showinfo("О программе", "Программа еще находится в разработке (70%)")
    
def openreadme(): 
    filename = './guide.txt'
    os.system("start " + filename) 

def quit_from_programm():
    sys.exit()

def create_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
            )
        print("Connection to MySQL DB successful")
        label_con = tk.Label(frame, text = "Соединение с базой данных успешно", 
                             font = ("Calibri", 12, "bold"), fg = "green")
        label_con.place(x=10, y=370, height=30, width=270)
    except Error as e:
        print(f"The error '{e}' occurred")
        label_con = tk.Label(frame, text = "Ошибка соединения с базой данных", 
                             font = ("Calibri", 12, "bold"), fg = "red")
        label_con.place(x=10, y=370, height=30, width=270)
    return connection


def open_change_win():
    if st_name == 0:
        messagebox.showerror("Ошибка!", "Для изменения необходимо выбрать ученика")
    else:
        s_change_name = StringVar()
        s_change_surname = StringVar()
        s_change_otchestvo = StringVar()
        pr_change_name = StringVar()
        pr_change_surname = StringVar()
        pr_change_otchestvo = StringVar()
        
        win_change = tk.Toplevel(win)
        win_change.title("INFO Change")
        win_change.iconbitmap('icon.ico')
        win_change.geometry("350x400+400+150")
        win_change.resizable(False,False)
        
        frame2 = tk.Frame(win_change, width=350, height=400)
        label2_1 = tk.Label(frame2, text = "Ученик", justify = 'left', font = ("Calibri", 14, "bold"))
        label2_2 = tk.Label(frame2, text = "Родитель", justify = 'left', font = ("Calibri", 14, "bold"))
        
        btnApply = tk.Button(frame2, text = "Применить изменения", font = ("Calibri", 12, "bold"), 
                      command = donothing)
        
        st_ent2_1 = tk.Entry(frame2, textvariable = s_change_name)
        st_ent2_2 = tk.Entry(frame2, textvariable = s_change_surname)
        st_ent2_3 = tk.Entry(frame2, textvariable = s_change_otchestvo)
        pr_ent2_1 = tk.Entry(frame2, textvariable = pr_change_name)
        pr_ent2_2 = tk.Entry(frame2, textvariable = pr_change_surname)
        pr_ent2_3 = tk.Entry(frame2, textvariable = pr_change_otchestvo)
        
        st_ent2_1.insert(0, st_name)
        st_ent2_2.insert(0, st_surname)
        st_ent2_3.insert(0, st_otchestvo)
        pr_ent2_1.insert(0, pr_name)
        pr_ent2_2.insert(0, pr_surname)
        pr_ent2_3.insert(0, pr_otchestvo)
        
        frame2.pack()
        label2_1.place(x=10, y=0, height=25, width=100)
        st_ent2_1.place(x=10, y=30, height=20, width=330)
        st_ent2_2.place(x=10, y=55, height=20, width=330)
        st_ent2_3.place(x=10, y=80, height=20, width=330)
        label2_2.place(x=10, y=110, height=25, width=100)
        pr_ent2_1.place(x=10, y=140, height=20, width=330)
        pr_ent2_2.place(x=10, y=165, height=20, width=330)
        pr_ent2_3.place(x=10, y=190, height=20, width=330)
        btnApply.place(x=20, y=360, height=30, width=310)
        
        
# НАСТРОЙКА ОКНА
win = tk.Tk()
win.title("AutoDoc")
win.iconbitmap('icon.ico')
win.geometry("700x420+400+150")
win.resizable(False,False)
win.minsize(300,350)
                                                                                
# МЕНЮ ПРОГРАММЫ
menuTop = tk.Menu(win)  
win.config(menu = menuTop)

submenuDown = tk.Menu(menuTop, tearoff = 0)
menuTop.add_cascade(label = 'Действие', menu = submenuDown)  
submenuDown.add_command(label = 'Составить договор', command = export) 
submenuDown.add_command(label = 'Список несортированных учеников', command = list_win_gui)
submenuDown.add_command(label = 'Импортировать анкету', command = import_ancet)
submenuDown.add_separator()
submenuDown.add_command(label = 'Выход из программы', command = quit_from_programm)

helpMenu = tk.Menu(menuTop, tearoff = 0)
menuTop.add_cascade(label = 'Справка', menu = helpMenu)
helpMenu.add_command(label = 'Иснтрукция по использованию', command = openreadme)
helpMenu.add_separator()
helpMenu.add_command(label = 'О программе', command = show_info_about_programm)

# ТЕЛО ПРОГРАММЫ
cnt_ancets = 0
st_name = 0
s_name = StringVar()
s_surname = StringVar()

frame = tk.Frame(win, width=700, height=400)
label1 = tk.Label(frame, text = "Имя ученика", font = ("Calibri", 15, "bold"))
label2 = tk.Label(frame, text = "Фамилия ученика", font = ("Calibri", 15, "bold"))
label4 = tk.Label(frame, text = "Информация", font = ("Calibri", 15, "bold"))
ent1 = tk.Entry(frame, textvariable = s_name)
ent2 = tk.Entry(frame, textvariable = s_surname)
textbox = tk.Text(frame)
btnSearch = tk.Button(frame, text = "Найти", font = ("Calibri", 12, "bold"), command = search_button)
btnContract = tk.Button(frame, text = "Сформировать договор", font = ("Calibri", 12, "bold"),
                        command = export)
btnDelete = tk.Button(frame, text = "Удалить ученика", font = ("Calibri", 12, "bold"), 
                      command = donothing)
btnChange = tk.Button(frame, text = "Изменить информацию", font = ("Calibri", 12, "bold"), 
                      command = open_change_win)

frame.pack()
label1.place(x=375, y=5, height=30, width=300)
label2.place(x=375, y=80, height=30, width=300)
label4.place(x=10, y=5, height=30, width=150)
ent1.place(x=375, y=40, height=30, width=300)
ent2.place(x=375, y=115 ,height=30, width=300)
textbox.place(x=10, y=35, height=330, width=330)
btnSearch.place(x=375, y=170, height=30, width=300)
btnContract.place(x=375, y=210, height=30, width=300)
btnChange.place(x=375, y=315, height=30, width=300)
btnDelete.place(x=375, y=355, height=30, width=300)

connection = create_connection("localhost", "root", "a$12df341212", "mydb")     
cursor = connection.cursor()
connection.commit()

# Разрываем подключение.
#connection.close()


# Цикл берет анкеты с сайта, удаляет их и перетаскивает в локальную папку (папку программы), 
# попутно уведомляя о новых анкетах

while True:
    try:
        dirname = 'C:/Apache24/htdocs/autodoc/files'
        dirfiles = os.listdir(dirname)
        
        url = 'http://127.0.0.1/autodoc/files/'+dirfiles[0]
        urllib.request.urlretrieve(url, f'./files/{dirfiles[0]}')
        file_path = f'C:/Apache24/htdocs/autodoc/files/{dirfiles[0]}'
        os.unlink(file_path)
        cnt_ancets += 1
    except:
        if cnt_ancets>0:
            messagebox.showinfo("Внимание!", "Появились новые заявки, для их рассмотрения: Действие>Список несортированных учеников")
        break
                                           
win.mainloop()